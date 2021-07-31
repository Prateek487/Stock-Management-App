import openpyxl
from datetime import datetime
from tkinter import *
from tkinter import messagebox
root = Tk()
root.title("Stock Master")
root.state("zoomed")
root.configure(background='lightgrey')
#Screen Module___________________________________________________________________________________________________
root.fullScreenState = False
root.attributes('-fullscreen',root.fullScreenState )

def toggleFullScreen(event):
   root.fullScreenState = not root.fullScreenState
   root.attributes("-fullscreen", root.fullScreenState)

def quitFullScreen(event):
   root.fullScreenState = False
   root.attributes("-fullscreen", root.fullScreenState)

root.bind("<F11>",toggleFullScreen)
root.bind("<Escape>",quitFullScreen)


#Frame1__________________________________________________________________________________________________________
frame1 = Frame(root)
frame1.pack(side = LEFT)
scrollbar = Scrollbar(root)
scrollbar.pack( side = RIGHT, fill = Y )
F1L1 = Label(frame1,text = "Name",font = ("Arial",20))
F1L2 = Label(frame1,text = "CFC",font = ("Arial",20))
F1L3 = Label(frame1,text = "Index",font = ("Arial",20))
F1L4 = Label(frame1,text = "Packs",font = ("Arial",20))
F1L5 = Label(frame1,text = "P/CFC",font = ("Arial",20))
F1L6 = Label(frame1,text = "MRP",font = ("Arial",20))
F1L7 = Label(frame1,text = "%Per",font = ("Arial",20))
Lb1 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 45)
Lb2 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 6)
Lb3 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 4)
Lb4 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 6)
Lb5 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 6)
Lb6 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 8)
Lb7 = Listbox(frame1,yscrollcommand = scrollbar.set, font = ("Arial",20),width = 6)
with open("txtName.txt","r+") as file1, open("txtData.txt","r+") as file2:
   lines1 = file1.readlines()
   lines2 = file2.readlines()
i = 0
for line1,line2 in zip(lines1,lines2):
   word = line2.split()
   Lb1.insert(i,line1)
   Lb2.insert(i,word[0])
   Lb4.insert(i,word[1])
   Lb5.insert(i,word[2])
   Lb3.insert(i,str(i+1) + ".")
   Lb6.insert(i,word[3])
   Lb7.insert(i,word[4])
   i += 1
F1L3.grid(row = 0,column = 0)
F1L1.grid(row = 0,column = 1)
F1L2.grid(row = 0,column = 2)
F1L4.grid(row = 0,column = 3)
F1L5.grid(row = 0,column = 4)
F1L6.grid(row = 0,column = 5)
F1L7.grid(row = 0,column = 6)
Lb3.grid(row = 1,column = 0)
Lb1.grid(row = 1,column = 1)
Lb2.grid(row = 1,column = 2)
Lb4.grid(row = 1,column = 3)
Lb5.grid(row = 1,column = 4)
Lb6.grid(row = 1,column = 5)
Lb7.grid(row = 1,column = 6)
def scroll_view(x,y):
   Lb2.yview(x,y)
   Lb1.yview(x,y)
   Lb3.yview(x,y)
   Lb4.yview(x,y)
   Lb5.yview(x,y)
   Lb6.yview(x,y)
   Lb7.yview(x,y)
scrollbar.config(command = scroll_view)
#Functions_______________________________________________________________________________________________________
def donothing():
   filewin = Toplevel(root)
   button = Button(filewin, text="Do nothing button",command = filewin.destroy)
   button.pack()
def onMouseWheel(event):
   w = event.widget
   if w is Lb1:
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb2:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb3:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb4:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb5:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb6:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
   elif w is Lb7:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
   else:
      Lb1.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb2.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb3.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb4.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb5.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb6.yview_scroll(int(-4*(event.delta/120)), "units")
      Lb7.yview_scroll(int(-4*(event.delta/120)), "units")
def AddtoList(a,b,c,d,e,f,window):
   if(a == ""):
      messagebox.showerror("Error","Name can't be empty")
      return
   if(b ==""):
      b = "0"
   elif(not(b.isnumeric())):
      messagebox.showerror("Error","Invalid Entry")
      return
   if(c == ""):
      c = "0"
   elif(not(c.isnumeric())):
      messagebox.showerror("Error","Invalid Entry")
      return
   if(d == ""):
      d = "1"
   elif(not(d.isnumeric())):
      messagebox.showerror("Error","Invalid Entry")
      return
   if(e == ""):
      e = "1"
   elif(not(e.replace('.', '', 1).isdigit())):
      messagebox.showerror("Error","Invalid Entry")
      return
   if(f == ""):
      f = "1"
   elif(not(f.replace('.', '', 1).isdigit())):
      messagebox.showerror("Error","Invalid Entry")
      return
   total = int(c) + int(b)*int(d)
   b = int(total%int(d))
   c = int(total/int(d))
   try:
      with open("txtName.txt","a") as file1, open("txtData.txt","a") as file2:
          file1.write(a + "\n")
          file2.write("%s %s %s %s %s\n" %(c,b,d,e,f))
      messagebox.showinfo("Done", "New item added succesfully")
      Lb1.insert(END,str(a))
      Lb2.insert(END,str(c))
      Lb4.insert(END,str(b))
      Lb5.insert(END,str(d))
      Lb6.insert(END,str(e))
      Lb7.insert(END,str(f))
      Lb3.insert(END,str(Lb3.size()+1) + ".")
   except:
      messagebox.showerror("Error","Error")
   window.destroy()
def Additem():                                                                                               
   filewin = Toplevel(root)
   filewin.geometry("530x100+400+150")
   filewin.transient(root)
   filewin.grab_set()
   L1 = Label(filewin,text = "Item Name")
   L1.grid(row = 0,column = 0)
   E1 = Entry(filewin)
   E1.grid(row = 0,column = 1)
   L2 = Label(filewin,text = "CFC")
   L2.grid(row = 0,column = 2)
   E2 = Entry(filewin)
   E2.grid(row = 0,column = 3)
   L3 = Label(filewin,text = "Packs")
   L3.grid(row = 0,column = 4)
   E3 = Entry(filewin)
   E3.grid(row = 0,column = 5)

   
   L4 = Label(filewin,text = "Packs per CFC")
   L4.grid(row = 1,column = 0)
   E4 = Entry(filewin)
   E4.grid(row = 1,column = 1)
   L5 = Label(filewin,text = "MRP")
   L5.grid(row = 1,column = 2)
   E5 = Entry(filewin)
   E5.grid(row = 1,column = 3)
   L6 = Label(filewin,text = "To be")
   L6.grid(row = 1,column = 4)
   E6 = Entry(filewin)
   E6.grid(row = 1,column = 5)
   button = Button(filewin, text="ADD ITEM" ,command = lambda:AddtoList(E1.get(),E2.get(),E3.get(),E4.get(),E5.get(),E6.get(),filewin))
   button.grid(row = 2,column = 3)

def DelfromList(window,str1,idx):
   x = 0
   y = -1
   if(idx == -1):
      for i in (Lb1.get(0,"end")):
         if(i.strip() == str1):
            idx = x
            break
         x += 1
   try:
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      i = 0
      with open ("txtName.txt","w") as f1,open ("txtData.txt","w") as f2:
         for line1,line2 in zip(lines1,lines2):
            if line1.strip("\n") != str1:
               f1.write(line1)
               f2.write(line2)
               
            else:
               if(idx != -1):
                  Lb1.delete(idx)
                  Lb2.delete(idx)
                  Lb4.delete(idx)
                  Lb5.delete(idx)
                  Lb6.delete(idx)
                  Lb7.delete(idx)
                  Lb3.delete(END)
               y = 0
               messagebox.showinfo("Done","' " + str1 + " '" + " Deleted successfully")
            i += 1
      if(y == -1):
         messagebox.showerror("Error","Item not found")
   except:
      messagebox.showerror("Error","Error")
   if(window):
      window.destroy()
def Deleteitem():
   filewin = Toplevel(root)
   filewin.transient(root)
   filewin.grab_set()
   filewin.geometry("400x50+400+300")
   L1 = Label(filewin,text = "Name of item to be deleted:")
   L1.grid(row = 0,column = 0)
   E1 = Entry(filewin)
   E1.grid(row = 0,column = 1)
   button = Button(filewin, text="Delete",command = lambda:DelfromList(filewin,E1.get(),-1))
   button.grid(row = 0,column = 2)

def Add_Del_peices(val,b,a,window,flag,idx):                                                             
   if(a ==""):
         a = "0"
   elif(not(a.isnumeric())):
         messagebox.showerror("Error","Invalid Entry")
         return
   if(b ==""):
         b = "0"
   elif(not(b.isnumeric())):
         messagebox.showerror("Error","Invalid Entry")
         return
   try:
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      x = flag*int(a)                        #Packs
      y = flag*int(b)                        #CFC
      i = 0
      for line1,line2 in zip(lines1,lines2):
         if line1.strip("\n") == val.strip("\n"):
            word = line2.split()
            x += int(word[1])
            y += int(word[0]) + int((x-x%(int(word[2])))/(int(word[2])))
            x = int(x%(int(word[2])))
            break
         i += 1
      if(x < 0 or y < 0):
         messagebox.showerror("Error","There is not enough stock")
         return
      lines2[i] = "%s %s %s %s %s\n" %(y,x,word[2],word[3],word[4])
      with open ("txtData.txt","w") as f2:
         f2.writelines(lines2)
      Lb2.insert(idx,y)
      Lb2.delete(idx+1)
      Lb4.insert(idx,x)
      Lb4.delete(idx+1)
      if flag == 1 :messagebox.showinfo("Done","%s Packs and %s CFC Added successfully" %(x,y))
      else :messagebox.showinfo("Done","%s Packs and %s CFC Removed successfully" %(x,y))
   except:
      messagebox.showerror("Error","Error")
   window.destroy()

def Add_Del_win(val,idx):
   filewin = Toplevel(root,background = "lightgrey")
   filewin.geometry("350x50+500+300")
   filewin.transient(frame1)
   filewin.grab_set()
   L1 = Label(filewin,text = "CFC",background = "lightgrey")
   L1.grid(row = 0,column = 0)
   E1 = Entry(filewin)
   E1.grid(row = 0,column = 1)
   L2 = Label(filewin,text = "Packs",background = "lightgrey")
   L2.grid(row = 0,column = 2)
   E2 = Entry(filewin)
   E2.grid(row = 0,column = 3)
   button1 = Button(filewin, text="Add",command = lambda:Add_Del_peices(val,E1.get(),E2.get(),filewin,1,idx))
   button1.grid(row = 1,column = 1)
   button2 = Button(filewin, text="Remove",command = lambda:Add_Del_peices(val,E1.get(),E2.get(),filewin,-1,idx))
   button2.grid(row = 1,column = 2)
   
def Modify(val,a,window,idx):
   if(a ==""):
         a = "0"
   elif((not(a.isnumeric())) or (a == "0")):
         messagebox.showerror("Error","Invalid Entry")
         return
   try:
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      i=0
      temp = ""
      cfc = 0
      packs = 0
      for line1,line2 in zip(lines1,lines2):
         if line1.strip("\n") == val.strip("\n"):
            word = line2.split()
            x = int(word[0])
            y = int(word[1])
            z = int(word[2])
            total = y + x*z
            if(int(a)>total):
               packs = str(total)
               cfc = 0
            else:
               packs = (int(total%int(a)))
               cfc = (int(total/int(a)))
            temp = "%s %s %s %s %s\n" %(cfc,packs,a,word[3],word[4])
            break
         i+=1
      lines2[i] =  temp + ""
      with open ("txtData.txt","w") as f2:
         f2.writelines(lines2)
      Lb5.insert(idx,a)
      Lb5.delete(idx+1)
      Lb2.insert(idx,cfc)
      Lb2.delete(idx+1)
      Lb4.insert(idx,packs)
      Lb4.delete(idx+1)
      messagebox.showinfo("Done","Field Modified successfully")
   except:
      messagebox.showerror("Error","Error")
   window.destroy()
def Modifym(val,a,window,idx,p):
   if(a ==""):
         a = "0"
   elif(not(a.replace('.', '', 1).isdigit())):
         messagebox.showerror("Error","Invalid Entry")
         return
   try:
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      i = 0
      for line1,line2 in zip(lines1,lines2):
         if line1.strip("\n") == val.strip("\n"):
            word = line2.split()
            if(int(p) == 6):
               word[3] = a;
            elif(int(p) == 7):
               word[4] = a;
            temp = "%s %s %s %s %s\n" %(word[0],word[1],word[2],word[3],word[4])
            break
         i+=1
      lines2[i] =  temp + ""
      with open ("txtData.txt","w") as f2:
         f2.writelines(lines2)
      if(int(p) == 6):
         Lb6.insert(idx,a)
         Lb6.delete(idx+1)
      elif(int(p) == 7):
         Lb7.insert(idx,a)
         Lb7.delete(idx+1)
      messagebox.showinfo("Done","Field Modified successfully")
   except:
      messagebox.showerror("Error","Error")
   window.destroy()

def ModifyCpP(val,index,p):
   filewin = Toplevel(root,background = "lightgrey")
   filewin.geometry("350x50+500+300")
   filewin.transient(frame1)
   filewin.grab_set()
   Lx = Label(filewin,text = "             ",background = "lightgrey")
   Lx.grid(row = 0,column = 0)
   L1 = Label(filewin,text = "Enter Pscks per CFC",background = "lightgrey")
   L1.grid(row = 0,column = 1)
   E1 = Entry(filewin)
   E1.grid(row = 0,column = 2)
   button1 = Button(filewin, text="Modify",command = lambda:Modify(val,E1.get(),filewin,index))
   button1.grid(row = 1,column = 2)
def ModifyMRP(val,index,p):
   filewin = Toplevel(root,background = "lightgrey")
   filewin.geometry("350x50+500+300")
   filewin.transient(frame1)
   filewin.grab_set()
   str1 = ""
   if(int(p) == 6):
      str1 = "Enter new MRP"
   elif(int(p) == 7):
      str1 = "Enter new Percentage"
   Lx = Label(filewin,text = "             ",background = "lightgrey")
   Lx.grid(row = 0,column = 0)
   L1 = Label(filewin,text = str1,background = "lightgrey")
   L1.grid(row = 0,column = 1)
   E1 = Entry(filewin)
   E1.grid(row = 0,column = 2)
   button1 = Button(filewin, text="Modify",command = lambda:Modifym(val,E1.get(),filewin,index,p))
   button1.grid(row = 1,column = 2)
def Search_box(a):
   if(a == ""):
      F2E1.delete(0,END)
   try:
      Lb1.delete(0,'end')
      Lb2.delete(0,'end')
      Lb3.delete(0,'end')
      Lb4.delete(0,'end')
      Lb5.delete(0,'end')
      Lb6.delete(0,'end')
      Lb7.delete(0,'end')
      i = 0
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      for line1,line2 in zip(lines1,lines2):
         word = line2.split()
         if (a in line1 or a.upper() in line1.upper()):
            Lb1.insert(i,line1)
            Lb2.insert(i,word[0])
            Lb4.insert(i,word[1])
            Lb5.insert(i,word[2])
            Lb6.insert(i,word[3])
            Lb7.insert(i,word[4])
            Lb3.insert(i,str(i+1) + ".")
            i += 1
      messagebox.showinfo("Search Results","Showing %d Results" % i)
   except:
      messagebox.showerror("Error","Something went Wrong")

def onselect(event,Listbox):
    w = event.widget
    try:
         idx = int(w.curselection()[0])
         val = str(Lb1.get(w.curselection()))
    except IndexError:
        return
    if Listbox is Lb1:
       return Add_Del_win(val,idx)
    elif Listbox is Lb5:
       return ModifyCpP(val,idx,5)
    elif Listbox is Lb6:
       return ModifyMRP(val,idx,6)
    elif Listbox is Lb7:
       return ModifyMRP(val,idx,7)
    return
def SaveToPrint():
   #try:
      wb = openpyxl.Workbook()
      sheet = wb.active
      c = []
      for i in range(0,7):
         c.insert(i,sheet.cell(row = 1, column = i+1))
      c[0].value = "Index"
      c[1].value = "Name"
      c[2].value = "CFC"
      c[3].value = "Packs"
      c[4].value = "Packs Per CFC"
      
      c[5].value = "MRP"
      c[6].value = "Percentage"
      with open ("txtName.txt","r") as f1,open ("txtData.txt","r") as f2:
         lines1 = f1.readlines()
         lines2 = f2.readlines()
      i = 2
      for line1,line2 in zip(lines1,lines2):
         ArrayVals = line2.split()
         k = []
         for j in range(1,8):
            k.insert(j-1,sheet.cell(row = i, column = j))
         k[0].value = i-1
         k[1].value = line1
         k[2].value = ArrayVals[0]
         k[3].value = ArrayVals[1]
         k[4].value = ArrayVals[2]
         
         k[5].value = ArrayVals[3]
         k[6].value = ArrayVals[4]
         i = i+1
      
      now = datetime.now()
      dt_string = now.strftime(" %d-%m-%Y %H-%M-%S")
      pq = "Public Documents"
      path = "C:\\Users\\Public\\" + pq +dt_string+".xlsx"
      wb.save(path)
      messagebox.showinfo("Done","File Saved Successfully at location: %s" %(path))
   #except:
    #  messagebox.showerror("Error","Something went wrong")
def About():
   filewin = Toplevel(root)
   filewin.title("Stock Master")
   filewin.geometry("845x180+300+100")
   filewin.resizable(0,0)
   txt ="\nFor source code and more related information please visit:\n"
   txt2 = "http://github.com/Prateek487"
   L2 = Label(filewin,text = txt,width = 60,font = ("Calibri",20))
   L1 = Entry(filewin,width = 60,font = ("Calibri",20),state = "readonly")
   var =  StringVar()
   var.set(txt2)
   L1.config(textvariable=var, relief='flat' , justify = "center")
   L1.config({"background" : "lightgrey"})
   L1.grid(row = 1,column = 0)
   L2.grid(row = 0,column = 0)
def Help():
   filewin = Toplevel(root)
   filewin.title("Stock Master")
   filewin.geometry("845x180+300+100")
   filewin.resizable(0,0)
   txt ="\nFor any help and query please contact me at :\n"
   txt2 = "Prateekjain487@gmail.com"
   L2 = Label(filewin,text = txt,width = 60,font = ("Calibri",20))
   L1 = Entry(filewin,width = 60,font = ("Calibri",20),state = "readonly")
   var =  StringVar()
   var.set(txt2)
   L1.config(textvariable=var, relief='flat' , justify = "center")
   L1.config({"background" : "lightgrey"})
   L1.grid(row = 1,column = 0)
   L2.grid(row = 0,column = 0)
def scrolldownby1():
   Lb5.yview_scroll(1,UNITS)
   Lb1.yview_scroll(1,UNITS)
   Lb2.yview_scroll(1,UNITS)
   Lb3.yview_scroll(1,UNITS)
   Lb4.yview_scroll(1,UNITS)
   Lb6.yview_scroll(1,UNITS)
   Lb7.yview_scroll(1,UNITS)
def scrollupby1():
   Lb5.yview_scroll(-1,UNITS)
   Lb1.yview_scroll(-1,UNITS)
   Lb2.yview_scroll(-1,UNITS)
   Lb3.yview_scroll(-1,UNITS)
   Lb4.yview_scroll(-1,UNITS)
   Lb6.yview_scroll(-1,UNITS)
   Lb7.yview_scroll(-1,UNITS)
def OnDown(event):
   w = event.widget
   idx = int(w.curselection()[0])
   if idx == w.size():
      return
   elif(w is Lb5):
      Lb5.select_clear(idx)
      Lb5.select_set(idx+1)
   elif(w is Lb4):
      Lb4.select_clear(idx)
      Lb4.select_set(idx+1)
   elif(w is Lb6):
      Lb6.select_clear(idx)
      Lb6.select_set(idx+1)
   elif(w is Lb7):
      Lb7.select_clear(idx)
      Lb7.select_set(idx+1)
   elif(w is Lb2):
      Lb2.select_clear(idx)
      Lb2.select_set(idx+1)
   elif(w is Lb1):
      Lb1.select_clear(idx)
      Lb1.select_set(idx+1)
   scrolldownby1()
def OnUp(event):
   w = event.widget
   idx = int(w.curselection()[0])
   if idx == 0:
      return
   elif(w is Lb6):
      Lb6.select_clear(idx)
      Lb6.select_set(idx-1)
   elif(w is Lb7):
      Lb7.select_clear(idx)
      Lb7.select_set(idx-1)
   elif(w is Lb5):
      Lb5.select_clear(idx)
      Lb5.select_set(idx-1)
   elif(w is Lb4):
      Lb4.select_clear(idx)
      Lb4.select_set(idx-1)
   elif(w is Lb2):
      Lb2.select_clear(idx)
      Lb2.select_set(idx-1)
   elif(w is Lb1):
      Lb1.select_clear(idx)
      Lb1.select_set(idx-1)
   scrollupby1()
#Search Bar___________________________________________________________________________________________________________
frame2 = Frame(root)
frame2.place(x = 120,y=0)

F2L1 = Label(frame2,text = "Enter Your Search:",font = ("Arial",20))
F2E1 = Entry(frame2,width = 30,font = ("Arial",20))
F2B1 = Button(frame2,text = "Search",command = lambda:Search_box(F2E1.get()),font = ("Arial",20))
F2B2 = Button(frame2,text = "Save To Print",command = lambda:SaveToPrint(),font = ("Arial",20))
F2B3 = Button(frame2,text = "Display All",command = lambda:Search_box(""),font = ("Arial",20))
F2L1.grid(row = 0,column = 1)
F2E1.grid(row = 0,column = 2)
F2B1.grid(row = 0,column = 3)
F2B2.grid(row = 0,column = 4)
F2B3.grid(row = 0,column = 0)

#Item Menu___________________________________________________________________________________________________________
menubar = Menu(root)
itemmenu = Menu(menubar, tearoff=0)
itemmenu.add_command(label="Add Item", command=Additem)
itemmenu.add_command(label="delete Item", command=Deleteitem)

itemmenu.add_separator()
menubar.add_cascade(label="Item", menu=itemmenu)

#Help Menu_____________________________________________________________________________________________________________
helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help", command=Help)
helpmenu.add_command(label="About...", command=About)
helpmenu.add_separator()
menubar.add_cascade(label="About", menu=helpmenu)
root.config(menu=menubar)

#Popup Menu_____________________________________________________________________________________________________________
def popup(event):
      try:
         w = event.widget
         idx = int(w.curselection()[0])
         str1 = str(w.get(w.curselection())).strip("\n")
         print(str1)
         try:
            root.popup_menu = Menu(root, tearoff=0)
            root.popup_menu.add_command(label="Delete",command=lambda :DelfromList("",str1,idx))
            root.popup_menu.tk_popup(event.x_root, event.y_root, 0)
         finally:
            root.popup_menu.grab_release()
      except:
         messagebox.showerror("Error","Please select an item first")

#Events__________________________________________________________________________________________________________________

Lb1.bind('<Double-1>', lambda e: onselect(e, Lb1))
Lb5.bind('<Double-1>', lambda e: onselect(e, Lb5))
Lb6.bind('<Double-1>', lambda e: onselect(e, Lb6))
Lb7.bind('<Double-1>', lambda e: onselect(e, Lb7))
Lb1.bind_all("<MouseWheel>", onMouseWheel)
Lb2.bind_all("<MouseWheel>", onMouseWheel)
Lb3.bind_all("<MouseWheel>", onMouseWheel)
Lb4.bind_all("<MouseWheel>", onMouseWheel)
Lb5.bind_all("<MouseWheel>", onMouseWheel)
Lb6.bind_all("<MouseWheel>", onMouseWheel)
Lb7.bind_all("<MouseWheel>", onMouseWheel)


Lb1.bind('<Button-3>',popup)
Lb1.bind('<Return>',lambda e: onselect(e, Lb1))
Lb5.bind('<Return>', lambda e: onselect(e, Lb5))
Lb6.bind('<Return>',lambda e: onselect(e, Lb6))
Lb7.bind('<Return>', lambda e: onselect(e, Lb7))
F2E1.bind('<Return>',lambda e:Search_box(F2E1.get()))

#left bind
Lb1.bind_all('<Down>', OnDown)
Lb2.bind_all('<Down>', OnDown)
Lb3.bind_all('<Down>', OnDown)
Lb4.bind_all('<Down>', OnDown)
Lb5.bind_all('<Down>', OnDown)
Lb6.bind_all('<Down>', OnDown)
Lb7.bind_all('<Down>', OnDown)

Lb1.bind_all('<Up>', OnUp)
Lb2.bind_all('<Up>', OnUp)
Lb3.bind_all('<Up>', OnUp)
Lb4.bind_all('<Up>', OnUp)
Lb5.bind_all('<Up>', OnUp)
Lb6.bind_all('<Up>', OnUp)
Lb7.bind_all('<Up>', OnUp)
root.mainloop()

#Main Loop_____________________________________________________________________________________________________________

